import sys
from PyQt5 import QtWidgets
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QFileDialog

import des
from openpyxl import load_workbook
from random import randint


class ExampleApp(QtWidgets.QMainWindow, des.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle('Definition Lookuper')
        self.setWindowIcon(QIcon('icon.ico'))
        self.setFixedSize(self.size())
        self.label.setText('')
        self.currentWord = ''
        self.wb = load_workbook(QFileDialog.getOpenFileName()[0])
        self.sheet = self.wb.worksheets[0]
        self.pickBtn.clicked.connect(self.pick)
        self.showDefBtn.clicked.connect(self.show_def)
        self.currInd = 0
        self.numWords = self.sheet.max_row

        self.actionChange_File.triggered.connect(self.changeFile)

        self.let1.clicked.connect(lambda: self.get_by_letter(self.let1.selectedItems()[0].text()))
        self.let2.clicked.connect(lambda: self.get_by_letter(self.let2.selectedItems()[0].text()))
        self.let3.clicked.connect(lambda: self.get_by_letter(self.let3.selectedItems()[0].text()))
        self.let4.clicked.connect(lambda: self.get_by_letter(self.let4.selectedItems()[0].text()))
        self.let5.clicked.connect(lambda: self.get_by_letter(self.let5.selectedItems()[0].text()))

        self.listWidget.clicked.connect(self.find)

    def pick(self):
        self.currInd = randint(1, self.sheet.max_row)
        cell = self.sheet.cell(row=self.currInd, column=1)
        while cell.value is None:
            self.currInd = randint(1, self.sheet.max_row)
            cell = self.sheet.cell(row=self.currInd, column=1)
        word = cell.value
        self.label.setText(word)

    def show_def(self):
        defn = self.sheet.cell(row=self.currInd, column=2).value
        self.textBrowser.setText(defn)

    def get_by_letter(self, letter):
        self.listWidget.clear()
        for cellObj in self.sheet['A1':f'A{self.numWords}']:
            for cell in cellObj:
                if cell.value is not None:
                    if cell.value[0].upper() == letter:
                        self.listWidget.addItem(cell.value)

    def find(self):
        value = self.listWidget.selectedItems()[0].text()
        for cellObj in self.sheet['A1':f'A{self.numWords}']:
            for cell in cellObj:
                if cell.value is not None:
                    if value.lower() == cell.value.lower():
                        self.label.setText(value)
                        self.currInd = cell.row
                        self.show_def()
                        return True

    def changeFile(self):
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFile)
        dialog.setNameFilter("Table (*.xlsx)")
        if dialog.exec_():
            file_name = dialog.selectedFiles()[0]
            self.wb = load_workbook(file_name)
            self.sheet = self.wb.worksheets[0]
            self.numWords = self.sheet.max_row


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = ExampleApp()
    window.show()
    app.exec_()


if __name__ == '__main__':
    main()
